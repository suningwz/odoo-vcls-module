from . import TranslatorOutlookContact
from . import ETL_SF
from . import generalSync
import openpyxl

import pytz
from simple_salesforce import Salesforce
from tzlocal import get_localzone
from datetime import datetime

from odoo import models, fields, api
import logging
_logger = logging.getLogger(__name__)

class OutlookContactSync(models.Model):
    _name = 'etl.outlook.contact'
    _inherit = 'etl.sync.mixin'

    def run(self, createInOdoo, updateInOdoo, nbMaxRecords):
        workbook = openpyxl.load_workbook(filename = r'C:\Users\laamouri\Desktop\CONTACT EV.xlsx')
        worksheet = workbook.active
        modifiedRecordsExt = list()
        
        for i in range(2, worksheet.max_row + 1):
            line = {}
            for j in range(1, worksheet.max_column + 1):
                line[worksheet.cell(row=1, column=j).value] = worksheet.cell(row=i, column=j).value
            modifiedRecordsExt.append(line)
            print(str(i) + '/' + str(worksheet.max_row + 1))

        translator = TranslatorOutlookContact.TranslatorOutlookContact()
        SF = self.env['etl.outlook.contact'].search([])
        if not SF:
            SF = self.env['etl.outlook.contact'].create({})

        SF.forceUpdateOdooInstance(translator,modifiedRecordsExt,None)

        """ ##### CODE HERE #####
        SF.updateKeyTable(modifiedRecordsExt)
        print('Updated key table done')
        if createInOdoo or updateInOdoo:
            SF.updateOdooInstance(translator,modifiedRecordsExt ,createInOdoo, updateInOdoo,nbMaxRecords)
        print('Updated odoo instance done')
        ##### CODE HERE ##### """
        SF.setNextRun() 
    
    def updateKeyTable(self, externalInstance):

        modifiedRecordsExt = externalInstance
        i = 0
        for extRecord in modifiedRecordsExt:
            try:
                self.toOdooId(extRecord['Id'])
            except (generalSync.KeyNotFoundError, ValueError):
                # Exist in External but not in Odoo
                self.addKeys(externalId = extRecord['Id'], odooId = None, state = 'needCreateOdoo')
                print('Update Key Table needCreateOdoo, ExternalId :{}'.format(extRecord['Id']))
                print(str(i) + '/' + str(len(modifiedRecordsExt)))
                i += 1


    def updateOdooInstance(self, translator,externalInstance, createInOdoo, updateInOdoo, nbMaxRecords):

        Modifiedrecords = externalInstance
        if not nbMaxRecords:
            nbMaxRecords = len(self.keys)
        i = 0
        for key in self.keys:
            item = None
            if i < nbMaxRecords:
                if key.state == 'needUpdateOdoo' and updateInOdoo:
                    for record in Modifiedrecords:
                        if record['Id'] == key.externalId:
                            item = record
                    if item:
                        try:
                            odooAttributes = translator.translateToOdoo(item, self, externalInstance)
                            record = self.env['res.partner'].search([('id','=',key.odooId)], limit=1)
                            record.write(odooAttributes)
                            print('Updated record in Odoo: {}'.format(item['First Name']))
                            _logger.info('Updated record in Odoo: {}'.format(item['First Name']))
                            key.state ='upToDate'
                            i += 1
                            print(str(i)+' / '+str(nbMaxRecords))
                        except ValueError as error:
                            _logger.error(error)
                elif key.state == 'needCreateOdoo' and createInOdoo:
                    for record in Modifiedrecords:
                        if record['Id'] == key.externalId:
                            item = record
                    if item:
                        odooAttributes = translator.translateToOdoo(item, self, externalInstance)
                        partner_id = self.env['res.partner'].create(odooAttributes).id
                        print('Create new record in Odoo: {}'.format(item['First Name']))
                        _logger.info('Create new record in Odoo: {}'.format(item['First Name']))
                        key.odooId = partner_id
                        key.state ='upToDate'
                        i += 1
                        print(str(i)+' / '+str(nbMaxRecords))
    
    def forceUpdateOdooInstance(self, translator,externalInstance,nbMaxRecords):
        print("force Update")
        Modifiedrecords = externalInstance
        if not nbMaxRecords:
            nbMaxRecords = len(self.keys)
        i = 0
        for item in Modifiedrecords:
            if i < nbMaxRecords:
                odooAttributes = translator.translateToOdoo(item, self, externalInstance)
                self.env['res.partner'].create(odooAttributes).id
                print('Create new record in Odoo: {}'.format(item['First Name']))
                _logger.info('Create new record in Odoo: {}'.format(item['First Name']))
                i += 1
                print(str(i) + ' / '+ str(nbMaxRecords))
                        